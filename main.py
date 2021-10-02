import openpyxl
import smtplib, ssl
from email.mime.text import MIMEText
from email.utils import formataddr
wb = openpyxl.load_workbook("V-EXPO REGISTRATION FORM (Responses).xlsx")

# fetching all the data from the given data sheet

sh1 = wb['Form Responses 1']      # sheet name
row = sh1.max_row                 # number of row in a sheet
column = sh1.max_column           # number of column in a sheet

data = []
name = []
for i in range(2, row+1):
    d = sh1.cell(i, 7).value
    data.append(d)
for i in range(2, row+1):
    d = sh1.cell(i, 2).value
    name.append(d)
print(data)
print(name)

#########################################################################
# User configuration
password = input("Enter your password:")
sender_email = "rockeytma@gmail.com"
sender_name = 'STICK Pybot'
receiver_emails = ['tmarakesh@gmail.com', 'moderntamilaringan@gmail.com ']
receiver_names = ['Rakesh', 'Sanjay']
port = 587
smtp_server = "smtp.gmail.com"
###########################################################################

# Email text
email_body = '''
    <!DOCTYPE html>
<html>
<head>
<meta name="viewport" content="width=device-width, initial-scale=1">
<style>
.container {
  position: relative;
  width: 100%;
  max-width: 400px;
}

.container img {
  width: 100%;
  height: auto;
}

.container .btn {
  position: absolute;
  top: 50%;
  left: 50%;
  transform: translate(-50%, -50%);
  -ms-transform: translate(-50%, -50%);
  background-color: #555;
  color: white;
  font-size: 16px;
  padding: 12px 24px;
  border: none;
  cursor: pointer;
  border-radius: 5px;
  text-align: center;
}

.container .btn:hover {
  background-color: black;
}
</style>
</head>
<body>
<div class="container">
  <img src="https://res.cloudinary.com/think-makers-asylum/image/upload/v1631690492/happy_jirrmk.jpg" alt="Happy engineering day" style="width:100%">

</div>
  <p>
    <h1> Why do we celebrate Engineer's Day ? </h1>
    Nation celebrates Engineer's Day on September 15 to commemorate the birth anniversary of the greatest Indian Engineer Bharat Ratna Mokshagundam Visvesvaraya. ... From 1912 to 1918, M. Visvesvaraya was appointed as the Diwan of Mysore. As a chief engineer, he constructed Mysore's famous Krishna Raja Sagara Dam.
  </p>
  <p><h3>Developed By Team Stick</h3></p>


</body>
</html>

'''
print("Sending the email...")
for receiver_email, receiver_name in zip(receiver_emails, receiver_names):
    # Configurating user's info
    msg = MIMEText(email_body, 'html')
    msg['To'] = formataddr((receiver_name, receiver_email))
    msg['From'] = formataddr((sender_name, sender_email))
    msg['Subject'] = 'Whats special today? Dear' + receiver_name

    try:
        # Creating a SMTP session | use 587 with TLS, 465 SSL and 25
        server = smtplib.SMTP('smtp.gmail.com', 587)
        # Encrypts the email
        context = ssl.create_default_context()
        server.starttls(context=context)
        # We log in into our Google account
        server.login(sender_email, password)
        # Sending email from sender, to receiver with the email body
        server.sendmail(sender_email, receiver_email, msg.as_string())
        print('Email sent!')
    except Exception as e:
        print(f'Oh no! Something bad happened!n {e}')
    finally:
        print('Closing the server...')
        server.quit()

